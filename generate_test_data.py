#!/usr/bin/env python3
"""
生成测试数据 — 创建三个示例表格用于测试脱敏程序。
运行: python generate_test_data.py
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── 模拟数据 ────────────────────────────────────────────────────────────

REAL_NAMES = [
    "张三丰", "李小明", "王建军", "赵丽华", "刘国强",
    "陈晓燕", "杨伟民", "黄桂英", "周志刚", "吴秀英",
    "郑浩然", "冯晓峰", "褚天宇", "卫文杰", "蒋鹏程",
]

EMP_IDS = [
    "E2024001", "E2024002", "E2024003", "E2024007", "E2024010",
    "EMP01234", "EMP05678", "EMP09012", "A21005", "B31008",
    "C41011", "D51015", "M2025-01", "M2025-02", "M2025-03",
]

ORGS = [
    "研发中心/后端开发组", "研发中心/前端开发组", "市场部/华东区",
    "市场部/华南区", "人力资源部/招聘组", "人力资源部/薪酬组",
    "财务部/会计核算", "财务部/预算管理", "运营部/客户服务",
    "产品部/移动端", "技术部/基础设施", "数据部/算法工程",
]

POSITIONS = [
    "高级工程师", "架构师", "产品经理", "市场总监",
    "人力资源专员", "财务分析师", "运营主管", "测试工程师",
    "数据科学家", "项目经理", "UI设计师", "技术VP",
]

PHONES = [
    "13800138001", "13912345678", "15088889999", "18601012345",
    "13500001111", "15822223333", "17777778888", "18966665555",
]

EMAILS = [
    "zhangsan@company.com", "lixm@company.com", "wangjj@company.com",
    "zhaolh@company.com", "liugq@company.com", "chenxy@company.com",
    "yangwm@company.com", "huanggy@company.com", "zhouzg@company.com",
    "wuxy@company.com",
]


def make_basic_table(wb, title_suffix=""):
    """创建基础员工信息表。"""
    ws = wb.active
    ws.title = f"员工信息{title_suffix}"

    headers = ["员工号", "姓名", "部门", "职位", "手机号", "邮箱"]
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, name in enumerate(REAL_NAMES):
        ws.append([
            EMP_IDS[i],
            name,
            random.choice(ORGS),
            random.choice(POSITIONS),
            random.choice(PHONES),
            EMAILS[i % len(EMAILS)],
        ])

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    return ws


def make_performance_table(wb):
    """创建绩效评估表（不同列名，测试模糊匹配）。"""
    ws = wb.create_sheet("绩效评估")

    headers = ["工号", "员工姓名", "所属组织", "绩效等级", "评分", "考核年度"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, name in enumerate(REAL_NAMES[:10]):
        ws.append([
            EMP_IDS[i],
            name,
            random.choice(ORGS),
            random.choice(["S", "A", "B+", "B", "C"]),
            round(random.uniform(60, 100), 1),
            2025,
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    return ws


def make_finance_table(wb):
    """创建薪酬统计表（更多列，含公式）。"""
    ws = wb.create_sheet("薪酬统计")

    headers = [
        "员工号", "姓名", "组织", "基本工资", "绩效工资",
        "津贴", "扣款", "实发工资"
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, name in enumerate(REAL_NAMES[:12]):
        base = random.randint(8000, 30000)
        perf = random.randint(2000, 8000)
        allowance = random.randint(500, 2000)
        deduct = random.randint(300, 2000)
        row_idx = i + 2
        # 写入值
        ws.cell(row=row_idx, column=1, value=EMP_IDS[i])
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=random.choice(ORGS))
        ws.cell(row=row_idx, column=4, value=base)
        ws.cell(row=row_idx, column=5, value=perf)
        ws.cell(row=row_idx, column=6, value=allowance)
        ws.cell(row=row_idx, column=7, value=deduct)
        # 实发工资 = 基本+绩效+津贴-扣款
        ws.cell(row=row_idx, column=8).value = f"=D{row_idx}+E{row_idx}+F{row_idx}-G{row_idx}"

    # 数值格式
    for col in range(4, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 16

    return ws


def get_column_letter(col_index):
    """Convert column index to letter (1→A, 2→B, ...)"""
    from openpyxl.utils import get_column_letter as _gcl
    return _gcl(col_index)


# ── 生成文件 ────────────────────────────────────────────────────────────

def main():
    # 文件一：员工信息表
    wb1 = Workbook()
    make_basic_table(wb1, "")
    # 删除默认空 sheet
    if "Sheet" in wb1.sheetnames and wb1.sheetnames[0] != "员工信息":
        del wb1["Sheet"]
    wb1.save("input/员工信息表.xlsx")
    print("✅ 生成: input/员工信息表.xlsx")

    # 文件二：绩效评估表（不同列名）
    wb2 = Workbook()
    make_performance_table(wb2)
    if "Sheet" in wb2.sheetnames:
        del wb2["Sheet"]
    wb2.save("input/绩效评估表.xlsx")
    print("✅ 生成: input/绩效评估表.xlsx")

    # 文件三：薪酬统计表（含公式）
    wb3 = Workbook()
    make_finance_table(wb3)
    if "Sheet" in wb3.sheetnames:
        del wb3["Sheet"]
    wb3.save("input/薪酬统计表.xlsx")
    print("✅ 生成: input/薪酬统计表.xlsx")

    print(f"\n🎉 共生成 3 个测试文件，分别使用不同的列名风格")
    print(f"   查看 input/ 目录确认，然后运行:")
    print(f"   cd /home/ubuntu/table-desensitizer")
    print(f"   python mask_tables.py")


if __name__ == "__main__":
    from openpyxl.utils import get_column_letter
    main()
