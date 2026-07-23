#!/usr/bin/env python3
"""
表格脱敏工具 — 批量处理 Excel 表格中的敏感信息
===============================================

支持对员工号、姓名、组织等字段进行多种策略的脱敏处理。
同一原始值始终映射为同一脱敏值（确定性映射），保证数据关联可用。

用法:
    python mask_tables.py                          # 使用默认 config.yaml
    python mask_tables.py --config my_config.yaml  # 指定配置文件
    python mask_tables.py --dry-run                # 预览模式，不实际写入
    python mask_tables.py --input-dir ./data       # 临时指定输入目录

依赖: openpyxl, pyyaml
"""

import argparse
import hashlib
import os
import random
import re
import sys
import time
from copy import copy
from pathlib import Path

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# =========================================================================
# 1. 配置结构
# =========================================================================

DEFAULT_CONFIG = {
    "input_dir": "./input",
    "output_dir": "./output",
    "suffix": "_masked",
    "header_row": 1,           # 表头行号（1 起），设为 "auto" 自动检测
    "rules": [
        {"column": "员工号", "strategy": "partial_mask",
         "options": {"prefix_len": 1, "suffix_len": 3}},
        {"column": "姓名", "strategy": "fake_chinese_name"},
        {"column": "组织", "strategy": "generic_org"},
    ],
    "encoding": "utf-8",
    "verbose": True,
}


def load_config(path="config.yaml"):
    """加载 YAML 配置文件，缺失字段用默认值补全。"""
    config = dict(DEFAULT_CONFIG)
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            user_config = yaml.safe_load(f) or {}
        config.update(user_config)
    else:
        print(f"⚠ 未找到配置文件 {path}，使用默认配置")
    return config


# =========================================================================
# 2. 确定性伪数据生成器
# =========================================================================

# 常见中文姓氏（百家姓前50）
SURNAMES = list("""
    赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨
    朱秦尤许何吕施张孔曹严华金魏陶姜
    戚谢邹喻柏水窦章云苏潘葛奚范彭郎
    鲁韦昌马苗凤花方俞任袁柳丰鲍史唐
    费廉岑薛雷贺倪汤滕殷罗毕郝邬安常
""".replace("\n", "").replace(" ", ""))

# 常用汉字（用于生成名字）
GIVEN_CHARS = list("""
    明华建国志文杰斌晓伟丽敏静霞婷雪燕
    磊强军平刚桂英芳萍红玲芬芳春辉浩宇
    轩宇辰昊博睿涵玥鹏程嘉怡泽洋瑞霖恺
""".replace("\n", "").replace(" ", "")) * 3  # 扩充池子

# 通用组织名称
ORG_PREFIXES = ["部门", "事业部", "中心", "处室", "科室", "项目组", "委员会"]
ORG_CATEGORIES = [
    "综合管理", "技术研发", "市场销售", "人力资源", "财务管理",
    "运营服务", "战略规划", "产品设计", "质量保障", "客户关系",
    "信息技术", "法务合规", "采购供应", "生产制造", "物流仓储",
]
ORG_TIERS = ["A", "B", "C", "D", "E", "F", "G"]


def _seed_from(value: str) -> int:
    """从字符串生成确定性种子。"""
    h = hashlib.md5(value.encode("utf-8")).hexdigest()
    return int(h[:8], 16)


def _rng(value: str):
    """基于原始值创建确定性随机数生成器。"""
    return random.Random(_seed_from(value))


def mask_partial(value: str, prefix_len=1, suffix_len=3, mask_char="*"):
    """
    部分掩码：保留前后部分，中间用掩码字符替换。
    例: "E2024001" → "E****001"  (prefix_len=1, suffix_len=3)
         "张三丰" → "张*丰"      (prefix_len=1, suffix_len=1)
    """
    if not value:
        return value
    s = str(value).strip()
    if len(s) <= prefix_len + suffix_len:
        # 太短则只保留首尾各1位
        if len(s) <= 2:
            return s[0] + mask_char * (len(s) - 1)
        return s[0] + mask_char * (len(s) - 2) + s[-1]
    return s[:prefix_len] + mask_char * (len(s) - prefix_len - suffix_len) + s[-suffix_len:]


def fake_chinese_name(value: str):
    """
    用伪中文姓名替换。同一原始值永远得到同一"假名"。
    假名由姓氏 + 1~2字名组成，风格接近真实姓名。
    """
    if not value or not str(value).strip():
        return value
    r = _rng(str(value).strip())
    surname = r.choice(SURNAMES)
    name_len = r.choices([1, 2], weights=[3, 7])[0]  # 70% 双字名
    given = "".join(r.choices(GIVEN_CHARS, k=name_len))
    return surname + given


def generic_org(value: str):
    """
    将组织名称替换为通用名称（如 "部门A-科室1"），
    同一原始组织名映射为同一通用名。
    """
    if not value or not str(value).strip():
        return value

    s = str(value).strip()
    r = _rng(s)

    # 根据原始组织的层级结构，生成对应层级的通用名
    parts = re.split(r'[/\\\-—·・、，,]', s)
    generic_parts = []
    for i, part in enumerate(parts):
        part = part.strip()
        if not part:
            continue
        tier = ORG_TIERS[i] if i < len(ORG_TIERS) else f"T{i}"
        # 每个层级从分类池中取一个
        cat = r.choice(ORG_CATEGORIES)
        generic_parts.append(f"{cat}-{tier}")
    return "/".join(generic_parts)


def hash_replace(value: str, length=8, prefix=""):
    """用哈希值前 N 位替换。"""
    if not value or not str(value).strip():
        return value
    s = str(value).strip()
    h = hashlib.sha256(s.encode("utf-8")).hexdigest()
    return f"{prefix}{h[:length]}"


def fake_employee_id(value: str):
    """
    伪员工号。保留原始格式特征（纯数字/字母数字混合），
    内容用确定性伪值替换。
    """
    if not value or not str(value).strip():
        return value
    s = str(value).strip()
    r = _rng(s)

    # 检测格式
    has_alpha = bool(re.search(r'[a-zA-Z]', s))
    is_all_digit = bool(re.fullmatch(r'\d+', s))

    if is_all_digit:
        # 纯数字：保留位数，生成伪数字
        length = len(s)
        return str(r.randint(10 ** (length - 1), 10**length - 1))
    elif has_alpha:
        # 字母数字混合：保留字母部分，数字随机
        prefix = re.match(r'^[a-zA-Z]+', s)
        suffix = re.search(r'\d+$', s)
        result = ""
        if prefix:
            # 保留字母前缀（或替换为随机字母）
            old_pre = prefix.group()
            new_pre = "".join(r.choices("ABCDEFGHJKLMNPQRSTUVWXYZ", k=len(old_pre)))
            result += new_pre
        if suffix:
            length = len(suffix.group())
            result += str(r.randint(10 ** (length - 1), 10**length - 1))
        if not result or len(result) < len(s):
            # fallback: 全替换
            result = "".join(r.choices("ABCDEFGHJKLMNPQRSTUVWXYZ0123456789", k=len(s)))
        return result

    # 兜底：哈希
    return hash_replace(s, length=min(len(s), 10))


# =========================================================================
# 3. 策略注册表
# =========================================================================

STRATEGIES = {
    "partial_mask": mask_partial,
    "fake_chinese_name": fake_chinese_name,
    "generic_org": generic_org,
    "hash": hash_replace,
    "fake_employee_id": fake_employee_id,
}


def apply_strategy(value, strategy_name, options=None):
    """根据策略名和选项对单个值进行脱敏。"""
    opts = options or {}
    func = STRATEGIES.get(strategy_name)
    if not func:
        raise ValueError(f"不支持的脱敏策略: {strategy_name}（支持: {list(STRATEGIES.keys())}）")

    # 把 options 传给函数
    sig = func.__code__.co_varnames[: func.__code__.co_kwonlyargcount + func.__code__.co_nlocals]
    kwargs = {k: v for k, v in opts.items() if k in sig}
    return func(value, **kwargs)


# =========================================================================
# 4. Excel 文件处理
# =========================================================================

def _auto_detect_header_row(ws, column_names, max_scan=15):
    """
    自动检测表头行。
    策略：扫描前 max_scan 行，分数最高的行作为表头。
    计分规则：
      - +1  每个非空文本单元格
      - +3  每个匹配配置列名的单元格（精确或子串匹配）
      - -5  该行全是空（跳过）
    """
    best_row = 1
    best_score = -1

    for row in ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row or max_scan),
                            values_only=False):
        r = row[0].row
        text_count = 0
        match_count = 0
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            text_count += 1
            for cn in column_names:
                if cn in s or s in cn:
                    match_count += 1

        if text_count == 0:
            continue  # 空行跳过

        score = text_count + match_count * 3
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def find_column_indices(ws, column_names, header_row=1):
    """
    在 worksheet 中查找匹配的列。
    支持：
      - header_row 为整数：精确指定表头行
      - header_row 为 "auto" 或 0：自动检测
      - 精确匹配："员工号"
      - 子串匹配："姓名" 匹配 "员工姓名"
    返回: {配置中的列名: (列索引, 匹配的表头)}
    """
    # 确定表头行号
    hr = header_row
    if hr == "auto" or (isinstance(hr, int) and hr < 1):
        hr = _auto_detect_header_row(ws, column_names)
        if hr != header_row and hr != 1:
            print(f"     自动检测表头行: 第 {hr} 行")

    headers = {}
    for row in ws.iter_rows(min_row=hr, max_row=hr):
        for cell in row:
            if cell.value is not None:
                headers[str(cell.value).strip()] = cell.column

    if not headers:
        print(f"     表头行（第 {hr} 行）为空，跳过")
        return {}

    result = {}
    used_columns = set()
    for col_name in column_names:
        if col_name in headers:
            col_idx = headers[col_name]
        else:
            matched = [h for h in headers if col_name in h]
            if len(matched) == 0:
                print(f"  ⚠ 未找到列 '{col_name}'，跳过")
                continue
            elif len(matched) > 1:
                print(f"  ⚠ 列名 '{col_name}' 匹配到多个: {matched}，跳过（请在 config 中使用精确列名）")
                continue
            col_idx = headers[matched[0]]

        if col_idx in used_columns:
            header_name = next((h for h, idx in headers.items() if idx == col_idx), str(col_idx))
            print(f"  ⚠ 列 '{header_name}' 已被脱敏跳过（多个规则匹配同一列）")
            continue

        used_columns.add(col_idx)
        matched_header = next(h for h, idx in headers.items() if idx == col_idx)
        result[col_name] = (col_idx, matched_header)

    return result


def get_cell_value(cell):
    """安全获取单元格值（字符串形式）。"""
    v = cell.value
    if v is None:
        return ""
    return str(v)


def _read_with_calamine(filepath):
    """
    使用 python-calamine（Rust 实现）读取 xlsx 数据。
    完全不解析样式，对 WPS 损坏文件最友好。
    返回 {sheet_name: [rows]}，每行是单元格值的列表。
    """
    import python_calamine

    cwb = python_calamine.load_workbook(filepath)
    result = {}
    for sheet_name in cwb.sheet_names:
        sheet = cwb.get_sheet_by_name(sheet_name)
        rows = list(sheet.iter_rows())
        result[sheet_name] = rows
    return result


def _rebuild_openpyxl(data_dict):
    """
    从 calamine 读取的数据重建 openpyxl Workbook。
    保留表头和数据，不保留原样式。
    """
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    first = True
    for sheet_name, rows in data_dict.items():
        if first:
            ws = default_ws
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    return wb


def _repair_xlsx_styles(filepath):
    """
    修复损坏的 styles.xml。
    1. 尝试保留原样式内容（用 lxml 容错解析）
    2. 若解析失败，生成一个包含 50 组默认条目的样式表，
       确保原文件任何样式索引都不越界
    返回 BytesIO 对象。
    """
    import io
    import zipfile
    import re

    buf = io.BytesIO()

    # 尝试从原文件提取并保留样式信息
    original_styles = None
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            if 'xl/styles.xml' in z.namelist():
                original_styles = z.read('xl/styles.xml')
    except Exception:
        pass

    if original_styles is not None:
        # 尝试用 lxml 容错解析原样式（处理 WPS 非标准输出）
        try:
            from lxml import etree
            root = etree.fromstring(original_styles)
            ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
            # 提取关键元素数量
            fonts_n = max(len(root.findall(f'{ns}fonts/{ns}font')), 1)
            fills_n = max(len(root.findall(f'{ns}fills/{ns}fill')), 2)
            borders_n = max(len(root.findall(f'{ns}borders/{ns}border')), 1)
            xfs_n = max(len(root.findall(f'{ns}cellXfs/{ns}xf')), 1)
            # 重写 count 属性（WPS 可能写错）
            for tag, count in [('fonts', fonts_n), ('fills', fills_n),
                               ('borders', borders_n), ('cellStyleXfs', 1),
                               ('cellXfs', xfs_n)]:
                el = root.find(f'{ns}{tag}')
                if el is not None:
                    el.set('count', str(count))
            repaired = etree.tostring(root, xml_declaration=True, encoding='UTF-8',
                                      standalone=True)
            with zipfile.ZipFile(filepath, 'r') as zin:
                with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        if item.filename == 'xl/styles.xml':
                            zout.writestr(item, repaired)
                        else:
                            zout.writestr(item, zin.read(item.filename))
            buf.seek(0)
            return buf
        except Exception:
            pass  # lxml 也失败，走 fallback

    # ── Fallback：生成 50 组默认样式 ──────────────────────────
    N = 50
    parts = []
    parts.append('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
    parts.append('<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">')

    # fonts
    parts.append(f'<fonts count="{N}">')
    for _ in range(N):
        parts.append('<font><sz val="11"/><name val="SimSun"/></font>')
    parts.append('</fonts>')

    # fills
    parts.append(f'<fills count="{N}">')
    for i in range(N):
        if i == 0:
            parts.append('<fill><patternFill patternType="none"/></fill>')
        else:
            parts.append('<fill><patternFill patternType="gray125"/></fill>')
    parts.append('</fills>')

    # borders
    parts.append(f'<borders count="{N}">')
    for _ in range(N):
        parts.append('<border><left/><right/><top/><bottom/><diagonal/></border>')
    parts.append('</borders>')

    # cellStyleXfs
    parts.append(f'<cellStyleXfs count="1">')
    parts.append('<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>')
    parts.append('</cellStyleXfs>')

    # cellXfs — 重要：每项引用不同索引覆盖所有组合
    parts.append(f'<cellXfs count="{N}">')
    for i in range(N):
        fid = min(i, N - 1)
        gid = min(i, N - 1)
        bid = min(i, N - 1)
        parts.append(f'<xf numFmtId="0" fontId="{fid}" fillId="{gid}" borderId="{bid}" xfId="0"/>')
    parts.append('</cellXfs>')

    parts.append('</styleSheet>')
    fallback_styles = '\n'.join(parts).encode('utf-8')

    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == 'xl/styles.xml':
                    zout.writestr(item, fallback_styles)
                else:
                    zout.writestr(item, zin.read(item.filename))
    buf.seek(0)
    return buf


def _safe_load_workbook(filepath):
    """
    多策略安全加载 Excel 文件。
    某些 WPS/非标准工具生成的文件样式损坏，需要降级加载。
    """
    strategies = [
        ("标准模式", {}),
        ("仅数据模式", {"data_only": True}),
        ("无宏模式", {"keep_vba": False}),
        ("数据+无宏", {"data_only": True, "keep_vba": False}),
        ("严格模式", {"data_only": True, "keep_vba": False, "rich_text": True}),
    ]
    last_error = None
    for name, kwargs in strategies:
        try:
            wb = load_workbook(filepath, **kwargs)
            if name != "标准模式":
                print(f"     ⚡ 使用降级加载: {name}")
            return wb
        except Exception as e:
            last_error = e
            continue

    # 终极手段：修复损坏的 styles.xml
    print(f"     🔧 尝试修复 styles.xml...")
    try:
        buf = _repair_xlsx_styles(filepath)
        wb = load_workbook(buf)
        buf.close()
        print(f"     ✅ 样式修复成功")
        return wb
    except Exception as e:
        last_error = e
        print(f"     ⚠ 样式修复失败: {e}")

    # 最终兜底：用 python-calamine 直接读数据，重建 workbook
    print(f"     📦 尝试 python-calamine 直接读取...")
    try:
        data = _read_with_calamine(filepath)
        wb = _rebuild_openpyxl(data)
        print(f"     ✅ calamine 读取成功 ({len(data)} 个工作表)")
        return wb
    except Exception as e:
        raise Exception(
            f"openpyxl 全部策略失败 + calamine 也失败: {e}\n"
            f"  文件可能已损坏，请尝试用 Excel/WPS 打开后另存为再重试。"
        )


def process_file(filepath, config, input_dir=None):
    """处理单个 Excel 文件，返回处理行数或 None（出错时）。"""
    rel_path = os.path.relpath(filepath, input_dir) if input_dir else os.path.basename(filepath)
    print(f"\n📄 处理: {rel_path}")

    try:
        wb = _safe_load_workbook(filepath)
    except Exception as e:
        print(f"  ❌ 无法打开文件（所有策略均失败）: {e}")
        return None

    rules = config.get("rules", [])
    header_row = config.get("header_row", 1)
    # 解析 header_row（支持 "auto" 和整数）
    if isinstance(header_row, str) and header_row.strip().lower() == "auto":
        hr_config = "auto"
    else:
        try:
            hr_config = int(header_row)
        except (ValueError, TypeError):
            hr_config = 1
    column_names = [r["column"] for r in rules]
    total_masked = 0
    total_rows = 0

    for ws in wb.worksheets:
        sheet_name = ws.title
        print(f"   ↳ 工作表: {sheet_name}")

        # 解析当前工作表的表头行
        hr = hr_config
        if hr == "auto":
            hr = _auto_detect_header_row(ws, column_names)
            if hr != 1:
                print(f"     自动检测表头行: 第 {hr} 行")

        col_map = find_column_indices(ws, column_names, header_row=hr)
        if not col_map:
            print(f"     跳过（未找到匹配列）")
            continue

        # 将列信息与规则关联
        col_rule_map = {}
        for rule in rules:
            cn = rule["column"]
            if cn in col_map:
                col_idx, matched_header = col_map[cn]
                col_rule_map[col_idx] = (rule["strategy"], rule.get("options", {}))

        # 逐行处理（表头之后开始）
        row_count = 0
        for row in ws.iter_rows(min_row=hr + 1, values_only=False):
            row_count += 1
            any_change = False
            for cell in row:
                if cell.column in col_rule_map:
                    strategy_name, opts = col_rule_map[cell.column]
                    original = get_cell_value(cell)
                    if not original:
                        continue

                    masked = apply_strategy(original, strategy_name, opts)
                    if masked != original:
                        # 保留原始格式（数字变字符串需要特殊处理）
                        if isinstance(cell.value, (int, float)):
                            cell.value = masked
                        else:
                            cell.value = masked
                        any_change = True
            if any_change:
                total_masked += 1

        total_rows += row_count
        print(f"     处理行数: {row_count}")

    # 生成输出路径（保留子目录结构）
    if input_dir:
        rel_dir = os.path.dirname(rel_path)
        out_dir = os.path.join(config["output_dir"], rel_dir) if rel_dir else config["output_dir"]
        os.makedirs(out_dir, exist_ok=True)
    else:
        out_dir = config["output_dir"]
    base, ext = os.path.splitext(os.path.basename(filepath))
    out_name = f"{base}{config.get('suffix', '_masked')}{ext}"
    out_path = os.path.join(out_dir, out_name)

    wb.save(out_path)
    wb.close()

    print(f"  ✅ 已保存: {out_path}")
    return total_rows, total_masked


# =========================================================================
# 5. 主流程
# =========================================================================

def main():
    parser = argparse.ArgumentParser(
        description="表格脱敏工具 — 批量处理 Excel 表格中的敏感信息",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python mask_tables.py                          # 使用默认配置
  python mask_tables.py --config my_rules.yaml   # 自定义配置
  python mask_tables.py --dry-run                # 只预览不写入
  python mask_tables.py --input-dir ./机密表格    # 指定输入目录（递归搜索子目录）
        """,
    )
    parser.add_argument("--config", default="config.yaml",
                        help="配置文件路径 (默认: config.yaml)")
    parser.add_argument("--input-dir", default=None,
                        help="输入目录（覆盖 config 中的 input_dir）")
    parser.add_argument("--output-dir", default=None,
                        help="输出目录（覆盖 config 中的 output_dir）")
    parser.add_argument("--dry-run", action="store_true",
                        help="预览模式：仅显示将要处理的文件，不实际写入")
    parser.add_argument("--suffix", default=None,
                        help="输出文件后缀（覆盖 config 中的 suffix）")
    args = parser.parse_args()

    # 加载配置
    config = load_config(args.config)

    if args.input_dir:
        config["input_dir"] = args.input_dir
    if args.output_dir:
        config["output_dir"] = args.output_dir
    if args.suffix:
        config["suffix"] = args.suffix

    input_dir = config["input_dir"]
    output_dir = config["output_dir"]

    # 确保目录存在
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    # 显示配置摘要
    print("=" * 60)
    print("📋 表格脱敏工具")
    print("=" * 60)
    print(f"  输入目录:  {os.path.abspath(input_dir)}")
    print(f"  输出目录:  {os.path.abspath(output_dir)}")
    print(f"  文件后缀:  {config.get('suffix', '_masked')}")
    hr_display = config.get('header_row', 1)
    print(f"  表头行:    {hr_display}")
    print(f"  脱敏规则:")
    for rule in config.get("rules", []):
        opts = rule.get("options", {})
        opt_str = f" ({opts})" if opts else ""
        print(f"    - 列 '{rule['column']}' → 策略: {rule['strategy']}{opt_str}")
    print(f"  模式:      {'🔍 预览 (dry-run)' if args.dry_run else '⚡ 正式执行'}")
    print("-" * 60)

    # 查找 Excel 文件（递归搜索子目录）
    patterns = ["*.xlsx", "*.xls"]
    files = []
    for p in patterns:
        files.extend(Path(input_dir).rglob(p))
    files = sorted(set(files))  # 去重

    if not files:
        print(f"⚠ 在 {input_dir}/ 下未找到 .xlsx/.xls 文件")
        print(f"  请将表格放入 {input_dir}/ 目录后重试")
        sys.exit(1)

    print(f"\n📂 发现 {len(files)} 个文件:")
    for f in files:
        rel = os.path.relpath(f, input_dir)
        print(f"   - {rel}")

    if args.dry_run:
        print("\n🔍 预览模式结束，未写入任何文件。")
        return

    # 开始处理
    start_time = time.time()
    success_count = 0
    total_rows_processed = 0
    total_rows_masked = 0
    failed_files = []

    for f in files:
        result = process_file(str(f), config, input_dir=input_dir)
        if result is None:
            failed_files.append(os.path.relpath(f, input_dir))
        else:
            rows_total, rows_masked = result
            success_count += 1
            total_rows_processed += rows_total
            total_rows_masked += rows_masked

    # 汇总
    elapsed = time.time() - start_time
    print("\n" + "=" * 60)
    print("📊 处理完成")
    print("=" * 60)
    print(f"  成功:      {success_count}/{len(files)} 个文件")
    print(f"  处理行数:  {total_rows_processed} 行")
    print(f"  脱敏操作:  {total_rows_masked} 行涉及脱敏")
    print(f"  耗时:      {elapsed:.2f} 秒")
    if failed_files:
        print(f"  失败:      {', '.join(failed_files)}")
    print(f"  输出目录:  {os.path.abspath(output_dir)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
